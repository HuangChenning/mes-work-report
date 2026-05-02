#!/usr/bin/env python3
"""
MES 实施计划报工 Excel 报告生成器。
生成两个 Excel 文件：
  1. 零报工实施计划 — 指定时间范围内合计工时为 0
  2. 极低报工实施计划 — 合计工时 > 0 且 < threshold
可选：--filter 生成过滤子集文件（如巡检零报工）。
用法：
  python generate_plan_excel.py --months 3 --threshold 20 --output-dir .
  python generate_plan_excel.py --months 3 --filter checkType=巡检 --filter title=巡检
"""

import argparse
import json
import os
import subprocess
import sys
from collections import defaultdict
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def run_mes(cmd: list[str]) -> dict | list:
    full_cmd = ["mes", "-o", "json"] + cmd
    result = subprocess.run(full_cmd, capture_output=True, text=True)
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        print(f"[ERROR] mes 输出无法解析: {result.stdout[:200]}", file=sys.stderr)
        return {}


def _extract_list(data) -> list[dict]:
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ("list", "operateCallBackObj", "data"):
            val = data.get(key)
            if isinstance(val, list):
                return val
            if isinstance(val, dict):
                inner = val.get("list")
                if isinstance(inner, list):
                    return inner
    return []


def fetch_all_active_plans() -> list[dict]:
    all_plans = []
    page = 1
    while True:
        data = run_mes([
            "plan", "list", "--status", "1",
            "--page", str(page), "--page-size", "200",
        ])
        items = _extract_list(data)
        if not items:
            break
        all_plans.extend(items)
        total = data.get("total", 0) if isinstance(data, dict) else len(items)
        if len(all_plans) >= total or len(items) < 200:
            break
        page += 1
    return all_plans


def fetch_all_teams() -> list[dict]:
    data = run_mes(["util", "list-teams"])
    return data if isinstance(data, list) else _extract_list(data)


def fetch_team_statistics(team_id: int, from_date: str, to_date: str) -> list[dict]:
    all_records = []
    page = 1
    while True:
        data = run_mes([
            "statistics", "list",
            "--team-id", str(team_id),
            "--from", from_date, "--to", to_date,
            "--page", str(page), "--page-size", "200",
        ])
        items = _extract_list(data)
        if not items:
            break
        all_records.extend(items)
        has_next = False
        if isinstance(data, dict):
            has_next = data.get("hasNextPage", False)
        if not has_next or len(items) < 200:
            break
        page += 1
    return all_records


def build_monthly_hours_map(teams: list[dict], from_date: str, to_date: str) -> dict[str, dict[str, float]]:
    acc_monthly: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for i, team in enumerate(teams):
        tid = team.get("id") or team.get("teamId")
        tname = team.get("name") or team.get("teamName") or str(tid)
        if not tid:
            continue
        print(f"       [{i+1}/{len(teams)}] {tname} (id={tid})...")
        records = fetch_team_statistics(tid, from_date, to_date)
        for r in records:
            acc_id = r.get("accId", "")
            if not acc_id:
                continue
            raw_date = r.get("taskDate") or r.get("start") or ""
            month_key = raw_date[:7]
            hours = r.get("taskTime") or 0
            if month_key and hours:
                acc_monthly[acc_id][month_key] += hours
        print(f"         {len(records)} 条记录")
    return dict(acc_monthly)


def get_months_list(from_date: date, to_date: date) -> list[str]:
    months = []
    cur = from_date.replace(day=1)
    end = to_date.replace(day=1)
    while cur <= end:
        months.append(cur.isoformat()[:7])
        if cur.month == 12:
            cur = cur.replace(year=cur.year + 1, month=1)
        else:
            cur = cur.replace(month=cur.month + 1)
    return months


def analyze_plans(plans: list[dict], monthly_map: dict, months: list[str], threshold: float):
    zero_plans = []
    low_plans = []
    for p in plans:
        acc_id = p.get("accId", "")
        month_hours = {}
        total_hours = 0.0
        for m in months:
            h = round(monthly_map.get(acc_id, {}).get(m, 0), 1)
            month_hours[m] = h
            total_hours += h

        idle_months = sum(1 for h in month_hours.values() if h == 0)
        plan_hours = sum((e.get("taskTime") or 0) for e in (p.get("executorList") or []))
        executors = "、".join(
            e.get("executorName", "") for e in (p.get("executorList") or []) if e.get("executorName")
        )

        row = {
            "planId": p.get("id"),
            "title": p.get("title", ""),
            "companyName": p.get("companyName") or "",
            "companyId": p.get("companyId") or "",
            "contractName": p.get("contractName") or "",
            "contractNum": p.get("contractNum") or "",
            "contractId": p.get("contractId") or "",
            "checkType": p.get("checkTypeDesc") or "",
            "status": p.get("statusDesc") or "进行中",
            "startDate": (p.get("startDate") or "")[:10],
            "endDate": (p.get("endDate") or "")[:10],
            "idleMonths": idle_months,
            "monthHours": month_hours,
            "totalHours": round(total_hours, 1),
            "executor": executors or "",
            "planHours": plan_hours,
            "createdBy": p.get("createdByName") or "",
            "createdTime": (p.get("createdTime") or "")[:10],
            "deliver": p.get("deliver") or "",
        }

        if total_hours == 0:
            zero_plans.append(row)
        elif total_hours < threshold:
            low_plans.append(row)

    zero_plans.sort(key=lambda x: (-x["idleMonths"], x["totalHours"]))
    low_plans.sort(key=lambda x: (-x["idleMonths"], x["totalHours"]))
    return zero_plans, low_plans


def apply_filters(plans: list[dict], filters: list[tuple[str, str]]) -> list[dict]:
    """过滤计划列表。filters 为 [(field, keyword), ...]，任一匹配即保留（OR 逻辑）。"""
    if not filters:
        return plans
    result = []
    for p in plans:
        for field, keyword in filters:
            val = p.get(field, "") or ""
            if keyword in val:
                result.append(p)
                break
    return result


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=11)
FILLS = {
    3: PatternFill("solid", fgColor="FFCCCC"),
    2: PatternFill("solid", fgColor="FFEACC"),
    1: PatternFill("solid", fgColor="FFFFFACC"),
}
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def write_excel(filepath: str, plans: list[dict], months: list[str], sheet_title: str,
                is_low: bool, desc_lines: list[str]):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    headers = [
        "序号", "实施计划ID", "实施计划名称", "客户名称", "客户ID",
        "合同名称", "合同编号", "合同ID", "计划类型", "状态",
        "开始日期", "结束日期", "空置月数",
    ]
    widths = [6, 10, 50, 25, 8, 40, 12, 8, 10, 10, 12, 12, 10]

    if is_low:
        headers.append("合计工时(h)")
        widths.append(12)

    month_labels = [f"{m.split('-')[1]}月工时(h)" for m in months]
    headers.extend(month_labels)
    widths.extend([12] * len(month_labels))

    headers.extend(["执行人", "计划工时(h)", "创建人", "创建时间", "交付物要求", "备注"])
    widths.extend([20, 12, 10, 12, 35, 20])

    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(c)].width = w

    for i, p in enumerate(plans, 1):
        col = 1
        row_data = [
            i, p["planId"], p["title"], p["companyName"], p["companyId"],
            p["contractName"], p["contractNum"], p["contractId"],
            p["checkType"], p["status"], p["startDate"], p["endDate"],
            p["idleMonths"],
        ]
        if is_low:
            row_data.append(p["totalHours"])

        row_data.extend(p["monthHours"][m] for m in months)
        row_data.extend([p["executor"], p["planHours"], p["createdBy"], p["createdTime"], p["deliver"], ""])

        fill = FILLS.get(3 if p["idleMonths"] >= 3 else p["idleMonths"])
        for val in row_data:
            cell = ws.cell(i + 1, col, val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            cell.alignment = LEFT_WRAP if col in (3, 6, 22) else CENTER
            col += 1

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "C2"

    ws2 = wb.create_sheet("说明")
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 60
    for idx, line in enumerate(desc_lines, 1):
        ws2.cell(idx, 1, line[0]).font = Font(name="微软雅黑", size=11, bold=True)
        ws2.cell(idx, 2, line[1]).font = DATA_FONT

    wb.save(filepath)
    print(f"[OK] {os.path.basename(filepath)} — {len(plans)} 条记录")


def parse_filter(f: str) -> tuple[str, str]:
    if "=" not in f:
        print(f"[WARN] 忽略无效 filter（缺少 =）：{f}", file=sys.stderr)
        return ("", "")
    field, keyword = f.split("=", 1)
    field_map = {
        "checkType": "checkType", "type": "checkType", "计划类型": "checkType",
        "title": "title", "名称": "title", "计划名称": "title",
        "companyName": "companyName", "客户": "companyName", "公司": "companyName",
        "contractName": "contractName", "合同": "contractName",
        "executor": "executor", "负责人": "executor", "执行人": "executor",
        "contractNum": "contractNum", "合同编号": "contractNum",
    }
    mapped = field_map.get(field.strip(), field.strip())
    return (mapped, keyword.strip())


def main():
    parser = argparse.ArgumentParser(description="生成实施计划报工 Excel 报告")
    parser.add_argument("--months", type=int, default=3, help="回溯月数（默认 3）")
    parser.add_argument("--from", dest="from_date", help="起始日期 YYYY-MM-DD")
    parser.add_argument("--to", dest="to_date", help="截止日期 YYYY-MM-DD（默认今天）")
    parser.add_argument("--threshold", type=float, default=20, help="极低报工工时阈值（默认 20h）")
    parser.add_argument("--filter", dest="filters", action="append", default=[],
                        help="过滤条件，格式 field=keyword，可多次指定（OR 逻辑）。"
                             "支持: checkType=巡检, title=巡检, companyName=银行, executor=张三, contractNum=00032 等")
    parser.add_argument("--output-dir", default=".", help="输出目录")
    args = parser.parse_args()

    to_date = date.fromisoformat(args.to_date) if args.to_date else date.today()
    if args.from_date:
        from_date = date.fromisoformat(args.from_date)
    else:
        m = to_date.month - args.months
        y = to_date.year
        while m <= 0:
            m += 12
            y -= 1
        from_date = date(y, m, 1)

    months = get_months_list(from_date, to_date)
    from_str = from_date.isoformat()
    to_str = to_date.isoformat()

    print(f"[INFO] 时间范围：{from_str} ~ {to_str}，共 {len(months)} 个月：{months}")
    print(f"[INFO] 极低报工阈值：< {args.threshold}h")
    if args.filters:
        print(f"[INFO] 过滤条件：{args.filters}")

    print("[INFO] 拉取全部进行中实施计划...")
    plans = fetch_all_active_plans()
    print(f"       共 {len(plans)} 个进行中计划")

    print("[INFO] 拉取全团队报工记录...")
    teams = fetch_all_teams()
    print(f"       共 {len(teams)} 个团队")
    monthly_map = build_monthly_hours_map(teams, from_str, to_str)
    print(f"       涉及 {len(monthly_map)} 个有报工的合同子项")

    print("[INFO] 分析计划报工情况...")
    zero_plans, low_plans = analyze_plans(plans, monthly_map, months, args.threshold)
    print(f"       零报工：{len(zero_plans)} 个")
    print(f"       极低报工（<{args.threshold}h）：{len(low_plans)} 个")

    os.makedirs(args.output_dir, exist_ok=True)
    range_label = f"{from_str}至{to_str}"
    month_range_desc = f"{from_str} ~ {to_str}，共 {len(months)} 个月"
    filter_zero = f"{month_range_desc}，合计报工工时为 0 小时"
    filter_low = f"{month_range_desc}，合计报工工时 > 0 且 < {args.threshold} 小时"

    write_excel(
        os.path.join(args.output_dir, f"零报工实施计划_{range_label}.xlsx"),
        zero_plans, months, "零报工实施计划", False,
        [("说明", ""), ("生成时间", to_str), ("数据范围", "进行中实施计划（未到期）"),
         ("筛选条件", filter_zero), ("空置月数", "时间范围内报工工时为 0 的月份数"),
         ("颜色说明", ""), ("", "红色背景 = 空置3个月及以上（高风险）"),
         ("", "橙色背景 = 空置2个月（中风险）"), ("", "黄色背景 = 空置1个月（低风险）"),
         ("注意", "月度列显示报工工时（小时）；执行人列的计划工时为 MES 系统设置的工时配额")],
    )

    write_excel(
        os.path.join(args.output_dir, f"极低报工实施计划_{range_label}.xlsx"),
        low_plans, months, "极低报工实施计划", True,
        [("说明", ""), ("生成时间", to_str), ("数据范围", "进行中实施计划（未到期）"),
         ("筛选条件", filter_low), ("空置月数", "时间范围内报工工时为 0 的月份数"),
         ("合计工时", f"{months[0]} ~ {months[-1]} 累计报工工时（小时）"),
         ("注意", "月度列显示报工工时（小时）；执行人列的计划工时为 MES 系统设置的工时配额")],
    )

    if args.filters:
        parsed = [parse_filter(f) for f in args.filters]
        parsed = [(f, k) for f, k in parsed if f]

        filtered_zero = apply_filters(zero_plans, parsed)
        if filtered_zero:
            filter_desc = "、".join(f'"{k}"' for _, k in parsed)
            filter_label = "、".join(f"{f}含{k}" for f, k in parsed)
            fname_suffix = "_".join(k for _, k in parsed)
            write_excel(
                os.path.join(args.output_dir, f"{fname_suffix}零报工实施计划_{range_label}.xlsx"),
                filtered_zero, months, f"{fname_suffix}零报工实施计划", False,
                [("说明", ""), ("生成时间", to_str), ("数据范围", "进行中实施计划（未到期）"),
                 ("筛选条件", f'{month_range_desc}，{filter_label}且合计报工工时为 0 小时'),
                 ("空置月数", "时间范围内报工工时为 0 的月份数"),
                 ("颜色说明", ""), ("", "红色背景 = 空置3个月及以上（高风险）"),
                 ("", "橙色背景 = 空置2个月（中风险）"), ("", "黄色背景 = 空置1个月（低风险）"),
                 ("注意", "月度列显示报工工时（小时）；执行人列的计划工时为 MES 系统设置的工时配额")],
            )

        filtered_low = apply_filters(low_plans, parsed)
        if filtered_low:
            fname_suffix = "_".join(k for _, k in parsed)
            write_excel(
                os.path.join(args.output_dir, f"{fname_suffix}极低报工实施计划_{range_label}.xlsx"),
                filtered_low, months, f"{fname_suffix}极低报工实施计划", True,
                [("说明", ""), ("生成时间", to_str), ("数据范围", "进行中实施计划（未到期）"),
                 ("筛选条件", f'{month_range_desc}，{filter_label}且合计报工工时 > 0 且 < {args.threshold} 小时'),
                 ("空置月数", "时间范围内报工工时为 0 的月份数"),
                 ("合计工时", f"{months[0]} ~ {months[-1]} 累计报工工时（小时）"),
                 ("注意", "月度列显示报工工时（小时）；执行人列的计划工时为 MES 系统设置的工物配额")],
            )


if __name__ == "__main__":
    main()
